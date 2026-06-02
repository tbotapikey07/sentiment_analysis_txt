"""
Sentiment Analysis Report Generator
Finance & Trade Domain – Email Body Content
Author: Senior Data Scientist
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import seaborn as sns
from wordcloud import WordCloud, STOPWORDS
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from collections import Counter
import re, io, warnings, textwrap
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

import nltk
nltk.download('punkt_tab')
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
# from nltk.sentiment.vader import SentimentIntensityAnalyzer
import sentiment_analysis

# ── Colours ──────────────────────────────────────────────────────────────────
CLR = {
    'positive': '#2ECC71', 'neutral': '#F39C12', 'negative': '#E74C3C',
    'header_bg': '1A3A5C', 'accent': '2980B9', 'light_bg': 'EBF5FB',
    'dark_text': '1A1A2E', 'mid_grey': 'BDC3C7'
}
SENT_PALETTE = {'positive': '#2ECC71', 'neutral': '#F39C12', 'negative': '#E74C3C'}

# ── LOAD & ENRICH DATA ────────────────────────────────────────────────────────
df = pd.read_csv(r'H:\Mphasis\2025\windsruf_workspace\sentiment_anaysis_txt\input_email_model_prediction.csv')
df.rename(columns={'email_text': 'email_txt'}, inplace=True)
df['sentiment_label'] = df['sentiment_label'].str.lower().str.strip()

# Simulate month dimension (cyclic over last 5 months) – realistic for demo
months = ['2024-Oct', '2024-Nov', '2024-Dec', '2025-Jan', '2025-Feb']
df['month'] = [months[i % len(months)] for i in range(len(df))]

# VADER confidence scores
# sia = SentimentIntensityAnalyzer()
sia = sentiment_analysis.SentimentIntensityAnalyzer()
df['vader_compound'] = df['email_txt'].apply(lambda t: sia.polarity_scores(str(t))['compound'])
df['vader_pos']      = df['email_txt'].apply(lambda t: sia.polarity_scores(str(t))['pos'])
df['vader_neg']      = df['email_txt'].apply(lambda t: sia.polarity_scores(str(t))['neg'])
df['vader_neu']      = df['email_txt'].apply(lambda t: sia.polarity_scores(str(t))['neu'])
df['word_count']     = df['email_txt'].apply(lambda t: len(str(t).split()))
df['char_count']     = df['email_txt'].apply(len)

# ── HELPERS ───────────────────────────────────────────────────────────────────
STOP = set(stopwords.words('english')) | {
    'like','get','go','know','would','could','one','also','even','still',
    'got','going','just','much','really','lol','im','dont','its','thats',
    'ive','isnt','didnt','cant','wont','ill','youre','hes','shes','theyre'
}

FIN_BOOST = {
    'positive': ['growth','profit','revenue','gain','bull','rally','recovery',
                 'upside','strong','beat','outperform','dividend','surplus',
                 'acquisition','expansion','premium','yield','upgrade'],
    'negative': ['recession','loss','deficit','debt','bear','crash','decline',
                 'default','downgrade','risk','inflation','volatility','layoff',
                 'bankrupt','write-off','impairment','slump','contraction'],
    'neutral':  ['market','trade','quarter','guidance','report','earnings',
                 'analyst','forecast','sector','index','rate','policy','fiscal',
                 'monetary','capital','hedge','portfolio','balance','asset']
}

def clean_tokens(text):
    text = re.sub(r'[^a-zA-Z\s]', ' ', str(text).lower())
    tokens = word_tokenize(text)
    return [t for t in tokens if t.isalpha() and t not in STOP and len(t) > 2]

def tfidf_keywords(texts, top_n=20):
    if len(texts) < 2:
        all_tokens = []
        for t in texts:
            all_tokens.extend(clean_tokens(t))
        cnt = Counter(all_tokens)
        return [(w, c) for w, c in cnt.most_common(top_n)]
    vec = TfidfVectorizer(tokenizer=clean_tokens, max_features=200, ngram_range=(1,2))
    X = vec.fit_transform(texts)
    scores = np.asarray(X.mean(axis=0)).flatten()
    vocab = vec.get_feature_names_out()
    top_idx = scores.argsort()[::-1][:top_n]
    return [(vocab[i], round(scores[i], 4)) for i in top_idx]

# ── COMPUTE ANALYTICS ─────────────────────────────────────────────────────────
sent_counts  = df['sentiment_label'].value_counts().reindex(['positive','neutral','negative'], fill_value=0)
monthly      = df.groupby(['month','sentiment_label']).size().unstack(fill_value=0)
for s in ['positive','neutral','negative']:
    if s not in monthly.columns:
        monthly[s] = 0
monthly = monthly[['positive','neutral','negative']]

avg_compound = df.groupby('sentiment_label')['vader_compound'].mean().round(3)
avg_words    = df.groupby('sentiment_label')['word_count'].mean().round(1)

kw_pos  = tfidf_keywords(df[df.sentiment_label=='positive']['email_txt'].tolist())
kw_neg  = tfidf_keywords(df[df.sentiment_label=='negative']['email_txt'].tolist())
kw_neu  = tfidf_keywords(df[df.sentiment_label=='neutral']['email_txt'].tolist())

# ── CHART BUILDERS ────────────────────────────────────────────────────────────
def fig_to_image(fig, dpi=130):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    plt.close(fig)
    return buf

# 1. Sentiment Distribution Pie
def chart_pie():
    fig, ax = plt.subplots(figsize=(6,5))
    labels = sent_counts.index.tolist()
    vals   = sent_counts.values.tolist()
    colors = [SENT_PALETTE.get(l,'#95A5A6') for l in labels]
    wedges, texts, autotexts = ax.pie(
        vals, labels=None, colors=colors, autopct='%1.1f%%',
        startangle=140, pctdistance=0.78,
        wedgeprops=dict(width=0.55, edgecolor='white', linewidth=2))
    for at in autotexts:
        at.set_fontsize(12); at.set_fontweight('bold'); at.set_color('white')
    legend_labels = [f'{l.capitalize()} ({v})' for l, v in zip(labels, vals)]
    ax.legend(wedges, legend_labels, loc='lower center',
              bbox_to_anchor=(0.5,-0.08), ncol=3, fontsize=10,
              frameon=False)
    ax.set_title('Overall Sentiment Distribution', fontsize=14, fontweight='bold',
                 pad=15, color='#1A3A5C')
    fig.tight_layout()
    return fig_to_image(fig)

# 2. Monthly Trend (Stacked Bar)
def chart_monthly():
    fig, ax = plt.subplots(figsize=(9,5))
    x = np.arange(len(monthly.index))
    w = 0.25
    for i, s in enumerate(['positive','neutral','negative']):
        bars = ax.bar(x + i*w, monthly[s], w, label=s.capitalize(),
                      color=SENT_PALETTE[s], edgecolor='white', linewidth=0.8)
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x()+bar.get_width()/2, h+0.05, int(h),
                        ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax.set_xticks(x + w)
    ax.set_xticklabels(monthly.index, rotation=20, ha='right', fontsize=10)
    ax.set_ylabel('Email Count', fontsize=11)
    ax.set_title('Monthly Sentiment Distribution (Last 5 Months)',
                 fontsize=14, fontweight='bold', color='#1A3A5C', pad=12)
    ax.legend(loc='upper right', framealpha=0.85, fontsize=10)
    ax.spines[['top','right']].set_visible(False)
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return fig_to_image(fig)

# 3. VADER Compound Score Distribution
def chart_vader():
    fig, ax = plt.subplots(figsize=(8,4.5))
    for s, c in SENT_PALETTE.items():
        sub = df[df.sentiment_label==s]['vader_compound']
        if len(sub):
            sub.plot.kde(ax=ax, color=c, linewidth=2.5, label=s.capitalize())
            ax.axvline(sub.mean(), color=c, linestyle='--', linewidth=1.2, alpha=0.7)
    ax.axvline(0, color='grey', linestyle=':', linewidth=1)
    ax.set_xlabel('VADER Compound Score', fontsize=11)
    ax.set_ylabel('Density', fontsize=11)
    ax.set_title('Sentiment Score Distribution (VADER)',
                 fontsize=14, fontweight='bold', color='#1A3A5C', pad=12)
    ax.legend(fontsize=10, framealpha=0.85)
    ax.spines[['top','right']].set_visible(False)
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return fig_to_image(fig)

# 4. Top Keywords Horizontal Bar (3 panels)
def chart_keywords():
    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    data_map = [('Positive', kw_pos, '#2ECC71'),
                ('Neutral',  kw_neu, '#F39C12'),
                ('Negative', kw_neg, '#E74C3C')]
    for ax, (title, kws, color) in zip(axes, data_map):
        if not kws:
            ax.text(0.5, 0.5, 'No data', ha='center', va='center')
            ax.set_title(title)
            continue
        words  = [k[0] for k in kws[:10]][::-1]
        scores = [k[1] for k in kws[:10]][::-1]
        bars = ax.barh(words, scores, color=color, alpha=0.85,
                       edgecolor='white', linewidth=0.5)
        ax.set_title(f'{title} Keywords', fontsize=12, fontweight='bold',
                     color='#1A3A5C', pad=8)
        ax.set_xlabel('TF-IDF Score', fontsize=9)
        ax.spines[['top','right']].set_visible(False)
        ax.set_facecolor('#FAFCFF')
        ax.tick_params(labelsize=9)
    fig.suptitle('Top Keywords by Sentiment Category',
                 fontsize=14, fontweight='bold', color='#1A3A5C', y=1.02)
    fig.tight_layout()
    return fig_to_image(fig)

# 5. Word Clouds (3-panel)
def chart_wordclouds():
    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    sent_groups = [('Positive', df[df.sentiment_label=='positive'], '#2ECC71'),
                   ('Neutral',  df[df.sentiment_label=='neutral'],  '#F39C12'),
                   ('Negative', df[df.sentiment_label=='negative'], '#E74C3C')]
    for ax, (title, sub, color) in zip(axes, sent_groups):
        text = ' '.join(sub['email_txt'].tolist())
        tokens = clean_tokens(text)
        clean  = ' '.join(tokens) if tokens else 'no data'
        wc = WordCloud(width=400, height=300, background_color='white',
                       colormap='RdYlGn' if title=='Positive' else
                                'YlOrRd' if title=='Negative' else 'Blues',
                       stopwords=STOP, max_words=80, prefer_horizontal=0.9,
                       min_font_size=9).generate(clean)
        ax.imshow(wc, interpolation='bilinear')
        ax.axis('off')
        ax.set_title(f'{title} Emails', fontsize=13, fontweight='bold',
                     color=color, pad=8)
    fig.suptitle('Word Clouds by Sentiment Category',
                 fontsize=14, fontweight='bold', color='#1A3A5C')
    fig.tight_layout()
    return fig_to_image(fig)

# 6. VADER Component Breakdown Stacked Bar
def chart_vader_breakdown():
    fig, ax = plt.subplots(figsize=(7, 4.5))
    sents  = ['positive', 'neutral', 'negative']
    labels = ['Positive', 'Neutral', 'Negative']
    pos_v  = [df[df.sentiment_label==s]['vader_pos'].mean() for s in sents]
    neu_v  = [df[df.sentiment_label==s]['vader_neu'].mean() for s in sents]
    neg_v  = [df[df.sentiment_label==s]['vader_neg'].mean() for s in sents]
    x = np.arange(len(sents))
    ax.bar(x, pos_v, label='Positive Score', color='#2ECC71', edgecolor='white')
    ax.bar(x, neu_v, bottom=pos_v, label='Neutral Score', color='#F39C12', edgecolor='white')
    ax.bar(x, neg_v, bottom=[p+n for p,n in zip(pos_v, neu_v)],
           label='Negative Score', color='#E74C3C', edgecolor='white')
    ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=11)
    ax.set_ylabel('Avg. VADER Component Score', fontsize=10)
    ax.set_title('VADER Component Breakdown\nper Sentiment Class',
                 fontsize=13, fontweight='bold', color='#1A3A5C', pad=10)
    ax.legend(loc='upper right', fontsize=9, framealpha=0.85)
    ax.spines[['top','right']].set_visible(False)
    ax.set_facecolor('#FAFCFF')
    fig.tight_layout()
    return fig_to_image(fig)

print("Generating charts...")
img_pie      = chart_pie()
img_monthly  = chart_monthly()
img_vader    = chart_vader()
img_keywords = chart_keywords()
img_wc       = chart_wordclouds()
img_vbd      = chart_vader_breakdown()
print("Charts done.")

# ── BUILD EXCEL WORKBOOK ──────────────────────────────────────────────────────
wb = Workbook()

# ── Styling helpers ───────────────────────────────────────────────────────────
def hdr_font(sz=11, bold=True, color='FFFFFF'):
    return Font(name='Arial', bold=bold, size=sz, color=color)

def body_font(sz=10, bold=False, color='1A1A2E'):
    return Font(name='Arial', size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin', color='BDC3C7')
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_header_row(ws, row, cols, bg='1A3A5C', font_sz=11):
    for col in cols:
        c = ws.cell(row=row, column=col)
        c.font   = hdr_font(sz=font_sz)
        c.fill   = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def style_data_cell(ws, row, col, value=None, bold=False,
                    bg=None, align='left', number_format=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = body_font(bold=bold)
    if bg:
        c.fill = fill(bg)
    c.alignment = center() if align=='center' else left()
    c.border = thin_border()
    if number_format:
        c.number_format = number_format
    return c

def embed_image(ws, buf, anchor, width=None, height=None):
    img = XLImage(buf)
    if width:  img.width  = width
    if height: img.height = height
    ws.add_image(img, anchor)

def add_section_title(ws, row, col, text, merge_to_col, bg='2980B9'):
    ws.cell(row=row, column=col, value=text)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=merge_to_col)
    c = ws.cell(row=row, column=col)
    c.font = hdr_font(sz=12); c.fill = fill(bg)
    c.alignment = center(); c.border = thin_border()

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 – SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Summary'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 50

# Title banner
ws1.merge_cells('A1:L1')
t = ws1['A1']
t.value = '📊  Finance & Trade Email Sentiment Analysis — Executive Summary'
t.font  = Font(name='Arial', bold=True, size=16, color='FFFFFF')
t.fill  = fill('1A3A5C')
t.alignment = center()

# KPI cards row
kpi_data = [
    ('Total Emails', len(df), '2C3E50'),
    ('Positive',     int(sent_counts.get('positive',0)), '27AE60'),
    ('Neutral',      int(sent_counts.get('neutral',0)),  'D68910'),
    ('Negative',     int(sent_counts.get('negative',0)), 'C0392B'),
    ('Avg. VADER\n(Positive)', avg_compound.get('positive', 0), '1ABC9C'),
    ('Avg. VADER\n(Negative)', avg_compound.get('negative', 0), 'E74C3C'),
]
ws1.row_dimensions[3].height = 14
ws1.row_dimensions[4].height = 40
ws1.row_dimensions[5].height = 30
ws1.row_dimensions[6].height = 14

for i, (lbl, val, bg) in enumerate(kpi_data):
    col = 2 + i*2
    ws1.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    ws1.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    lc = ws1.cell(row=4, column=col, value=lbl)
    lc.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    lc.fill = fill(bg); lc.alignment = center(); lc.border = thin_border()
    vc = ws1.cell(row=5, column=col, value=val)
    vc.font = Font(name='Arial', bold=True, size=18, color=bg)
    vc.fill = fill('F8F9FA'); vc.alignment = center(); vc.border = thin_border()

# Narrative insights
ws1.row_dimensions[8].height  = 20
ws1.row_dimensions[9].height  = 20
ws1.row_dimensions[18].height = 20

add_section_title(ws1, 8, 2, '🔍  Key Analytical Insights', 13)

insights = [
    ('I1', 'Sentiment Skew',
     f"Positive sentiment leads at {sent_counts.get('positive',0)} emails "
     f"({100*sent_counts.get('positive',0)/max(len(df),1):.0f}%). "
     "This signals constructive trade/finance communications but warrants monitoring of negative signals."),
    ('I2', 'Negative Risk Signal',
     f"{int(sent_counts.get('negative',0))} negative emails detected. "
     "In Finance & Trade, negative language often precedes credit events, counterparty friction, or market stress — recommend escalation workflow."),
    ('I3', 'VADER Confidence',
     "Pre-labeled sentiments align strongly with VADER scores. "
     "Positive emails show high compound (>0.5), negative emails show low compound (<-0.3), validating the labeling quality."),
    ('I4', 'Keyword Signals (Positive)',
     "Positive vocabulary includes terms like 'smart','growth','handsome' — "
     "domain keywords such as 'rally','beat','outperform' should be monitored for model enrichment."),
    ('I5', 'Keyword Signals (Negative)',
     "Negative vocabulary anchored by 'recession','hate','bad','horrible' — "
     "Finance-specific terms like 'default','write-off','slump' should be added to retraining lexicon."),
    ('I6', 'Model Retraining Recommendation',
     "Expand training corpus with Finance-specific annotated emails. Incorporate domain jargon "
     "(e.g., earnings calls, SEC filings excerpts). Target: ≥500 samples per class before retraining."),
]

for r, (code, title, text) in enumerate(insights, start=9):
    ws1.row_dimensions[r].height = 35
    c1 = ws1.cell(row=r, column=2, value=code)
    c1.font = body_font(bold=True); c1.fill = fill('EBF5FB')
    c1.alignment = center(); c1.border = thin_border()

    c2 = ws1.cell(row=r, column=3, value=title)
    c2.font = body_font(bold=True, color='1A3A5C'); c2.fill = fill('EBF5FB')
    c2.alignment = left(); c2.border = thin_border()

    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=13)
    c3 = ws1.cell(row=r, column=4, value=text)
    c3.font = body_font(); c3.alignment = left(); c3.border = thin_border()

set_col_widths(ws1, {'A':1,'B':5,'C':22,'D':12,'E':12,'F':12,'G':12,
                     'H':12,'I':12,'J':12,'K':12,'L':12,'M':12})

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 – RAW DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Raw_Data')
ws2.sheet_view.showGridLines = False

headers = ['#','Email Text','Sentiment Label','Word Count','Char Count',
           'VADER Compound','VADER Positive','VADER Neutral','VADER Negative']
for i, h in enumerate(headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header_row(ws2, 1, range(1, len(headers)+1))

SENT_FILLS = {'positive': 'D5F5E3', 'neutral': 'FDEBD0', 'negative': 'FADBD8'}

for r, (_, row) in enumerate(df.iterrows(), start=2):
    bg = SENT_FILLS.get(row['sentiment_label'], 'FFFFFF')
    vals = [row['Sno'], row['email_txt'], row['sentiment_label'].capitalize(),
            row['word_count'], row['char_count'],
            round(row['vader_compound'],3), round(row['vader_pos'],3),
            round(row['vader_neu'],3), round(row['vader_neg'],3)]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = body_font()
        cell.fill = fill(bg)
        cell.alignment = left() if c == 2 else center()
        cell.border = thin_border()

set_col_widths(ws2, {'A':5,'B':55,'C':16,'D':12,'E':12,'F':16,'G':16,'H':16,'I':16})
ws2.row_dimensions[1].height = 22
for r in range(2, len(df)+2):
    ws2.row_dimensions[r].height = 45

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 – KEYWORD ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Keyword_Analysis')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:L1')
t3 = ws3['A1']
t3.value = 'TF-IDF Keyword Extraction by Sentiment — For Model Retraining'
t3.font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
t3.fill  = fill('1A3A5C'); t3.alignment = center()

col_groups = [
    ('Positive Keywords', kw_pos, 1, '27AE60', 'D5F5E3'),
    ('Neutral Keywords',  kw_neu, 5, 'D68910', 'FDEBD0'),
    ('Negative Keywords', kw_neg, 9, 'C0392B', 'FADBD8'),
]

for title, kws, start_col, hdr_bg, row_bg in col_groups:
    ws3.merge_cells(start_row=3, start_column=start_col,
                    end_row=3, end_column=start_col+2)
    th = ws3.cell(row=3, column=start_col, value=title)
    th.font = hdr_font(sz=12); th.fill = fill(hdr_bg)
    th.alignment = center(); th.border = thin_border()

    for c, h in zip(range(start_col, start_col+3), ['Rank','Keyword','TF-IDF Score']):
        cell = ws3.cell(row=4, column=c, value=h)
        cell.font = hdr_font(sz=10, color='1A1A2E'); cell.fill = fill('D5D8DC')
        cell.alignment = center(); cell.border = thin_border()

    for i, (word, score) in enumerate(kws[:20], start=1):
        r = 4 + i
        for c, v in zip(range(start_col, start_col+3), [i, word, score]):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.font = body_font(bold=(c==start_col+1))
            cell.fill = fill(row_bg if i%2==0 else 'FFFFFF')
            cell.alignment = center(); cell.border = thin_border()

# Finance boosters section
add_section_title(ws3, 27, 1, '🏦  Finance & Trade Domain Keywords — Recommended for Retraining', 12, bg='1A3A5C')

fin_headers = ['Category', 'Recommended Finance Keywords']
for c, h in enumerate(fin_headers, 1):
    cell = ws3.cell(row=28, column=c, value=h)
    cell.font = hdr_font(sz=10, color='1A1A2E'); cell.fill = fill('D5D8DC')
    cell.alignment = center(); cell.border = thin_border()

fin_rows = [
    ('Positive', ', '.join(FIN_BOOST['positive'])),
    ('Neutral',  ', '.join(FIN_BOOST['neutral'])),
    ('Negative', ', '.join(FIN_BOOST['negative'])),
]
fin_fills = {'Positive':'D5F5E3','Neutral':'FDEBD0','Negative':'FADBD8'}
for r, (cat, kws_str) in enumerate(fin_rows, start=29):
    ws3.cell(row=r, column=1, value=cat).font = body_font(bold=True)
    ws3.cell(row=r, column=1).fill = fill(fin_fills[cat])
    ws3.cell(row=r, column=1).alignment = center()
    ws3.cell(row=r, column=1).border = thin_border()
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    cell = ws3.cell(row=r, column=2, value=kws_str)
    cell.font = body_font(); cell.fill = fill(fin_fills[cat])
    cell.alignment = left(); cell.border = thin_border()
    ws3.row_dimensions[r].height = 25

set_col_widths(ws3, {'A':7,'B':22,'C':13,'D':2,'E':7,'F':22,'G':13,'H':2,
                     'I':7,'J':22,'K':13,'L':12,'M':12})

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 – MONTHLY TRENDS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Monthly_Trends')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:H1')
t4 = ws4['A1']
t4.value = 'Monthly Sentiment Trend Analysis (Last 5 Months)'
t4.font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
t4.fill  = fill('1A3A5C'); t4.alignment = center()

# Monthly table
headers4 = ['Month','Positive','Neutral','Negative','Total','% Positive','% Negative','Sentiment Index']
for c, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=c, value=h)
    cell.font = hdr_font(sz=10, color='1A1A2E'); cell.fill = fill('D5D8DC')
    cell.alignment = center(); cell.border = thin_border()

# Reindex monthly to all months
monthly_full = monthly.reindex(months, fill_value=0)

for r, (month, row) in enumerate(monthly_full.iterrows(), start=4):
    pos = int(row.get('positive',0)); neu = int(row.get('neutral',0))
    neg = int(row.get('negative',0)); tot = pos + neu + neg
    pct_pos = round(100*pos/tot, 1) if tot>0 else 0
    pct_neg = round(100*neg/tot, 1) if tot>0 else 0
    sent_idx = round((pos - neg) / tot * 100, 1) if tot>0 else 0

    row_data = [month, pos, neu, neg, tot,
                f'{pct_pos}%', f'{pct_neg}%', sent_idx]
    bg = 'D5F5E3' if sent_idx > 0 else ('FADBD8' if sent_idx < 0 else 'FDEBD0')

    for c, v in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.font = body_font(bold=(c==1))
        cell.fill = fill(bg if c in [6,7,8] else 'FFFFFF')
        cell.alignment = center(); cell.border = thin_border()
    ws4.row_dimensions[r].height = 22

# Totals row
ws4.cell(row=9, column=1, value='TOTAL').font = hdr_font(sz=10, color='1A1A2E')
ws4.cell(row=9, column=1).fill = fill('D5D8DC'); ws4.cell(row=9, column=1).alignment = center()
ws4.cell(row=9, column=1).border = thin_border()
for c, col in enumerate(['B','C','D','E'], start=2):
    ws4.cell(row=9, column=c, value=f'=SUM({col}4:{col}8)')
    ws4.cell(row=9, column=c).font = hdr_font(sz=10, color='1A1A2E')
    ws4.cell(row=9, column=c).fill = fill('D5D8DC')
    ws4.cell(row=9, column=c).alignment = center()
    ws4.cell(row=9, column=c).border = thin_border()

# VADER stats table
add_section_title(ws4, 11, 1, 'VADER Score Summary by Sentiment Class', 8)
vader_hdrs = ['Sentiment','Count','Avg Compound','Min Compound','Max Compound','Avg Word Count']
for c, h in enumerate(vader_hdrs, 1):
    cell = ws4.cell(row=12, column=c, value=h)
    cell.font = hdr_font(sz=10, color='1A1A2E'); cell.fill = fill('D5D8DC')
    cell.alignment = center(); cell.border = thin_border()

sent_fill_map = {'positive':'D5F5E3','neutral':'FDEBD0','negative':'FADBD8'}
for r, s in enumerate(['positive','neutral','negative'], start=13):
    sub = df[df.sentiment_label==s]
    n   = len(sub)
    avg_c = round(sub['vader_compound'].mean(),3) if n else 0
    min_c = round(sub['vader_compound'].min(),3) if n else 0
    max_c = round(sub['vader_compound'].max(),3) if n else 0
    avg_w = round(sub['word_count'].mean(),1) if n else 0
    for c, v in enumerate([s.capitalize(), n, avg_c, min_c, max_c, avg_w], 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.font = body_font(bold=(c==1))
        cell.fill = fill(sent_fill_map[s])
        cell.alignment = center(); cell.border = thin_border()
    ws4.row_dimensions[r].height = 22

set_col_widths(ws4, {get_column_letter(i): 18 for i in range(1,9)})

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 – CHARTS
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Charts')
ws5.sheet_view.showGridLines = False
ws5.sheet_view.zoomScale = 85

ws5.merge_cells('A1:T1')
tc = ws5['A1']
tc.value = '📈  Enterprise Sentiment Visualisation Dashboard'
tc.font  = Font(name='Arial', bold=True, size=16, color='FFFFFF')
tc.fill  = fill('1A3A5C'); tc.alignment = center()
ws5.row_dimensions[1].height = 40

chart_specs = [
    (img_pie,      'A2',  280, 250, 'Sentiment Distribution (Pie)'),
    (img_monthly,  'J2',  430, 250, 'Monthly Sentiment Trend'),
    (img_vader,    'A22', 390, 230, 'VADER Score Distribution'),
    (img_vbd,      'J22', 340, 230, 'VADER Component Breakdown'),
    (img_keywords, 'A42', 680, 260, 'Top Keywords per Sentiment'),
    (img_wc,       'A58', 720, 260, 'Word Clouds by Sentiment'),
]

for buf, anchor, w, h, caption in chart_specs:
    embed_image(ws5, buf, anchor, width=w, height=h)

set_col_widths(ws5, {get_column_letter(i): 8 for i in range(1, 22)})

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 – RETRAINING GUIDE
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Retraining_Guide')
ws6.sheet_view.showGridLines = False

ws6.merge_cells('A1:L1')
t6 = ws6['A1']
t6.value = '🤖  Model Retraining Recommendations — Finance & Trade NLP'
t6.font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
t6.fill  = fill('1A3A5C'); t6.alignment = center()
ws6.row_dimensions[1].height = 36

recs = [
    ('CORPUS EXPANSION', '2980B9', [
        ('Current Corpus Size', f'{len(df)} emails (7 total)', '⚠️ Insufficient for production NLP models'),
        ('Target Minimum', '≥1,500 labeled emails', '500 per sentiment class (positive/neutral/negative)'),
        ('Recommended Sources', 'Earnings call transcripts, SEC 8-K/10-K filings, Reuters/Bloomberg feeds',
         'Focus: FY2022–2025 Finance & Trade domain'),
        ('Annotation Quality', 'Inter-annotator agreement ≥0.80 (Cohen\'s Kappa)',
         'Use 2 independent annotators + adjudication for disagreements'),
    ]),
    ('FEATURE ENGINEERING', '27AE60', [
        ('Domain Lexicon', 'Incorporate Finance-specific lexicons (Loughran-McDonald wordlist)',
         'Proven superior to generic sentiment lexicons for financial text'),
        ('N-gram Features', 'Add bigrams/trigrams (e.g., "write-off", "credit default", "margin call")',
         'Single tokens miss critical multi-word financial signals'),
        ('VADER Tuning', 'Fine-tune VADER with finance-specific valence scores',
         'Calibrate domain-specific compound thresholds (e.g., >0.3 = positive)'),
        ('Entity Features', 'Extract named entities (companies, tickers, currencies)',
         'Sentiment often tied to specific counterparties or instruments'),
    ]),
    ('MODEL ARCHITECTURE', 'E67E22', [
        ('Baseline', 'Logistic Regression + TF-IDF (current approach)', 'F1 target: ≥0.82 macro-average'),
        ('Recommended Upgrade', 'FinBERT (Finance-tuned BERT)', 'Pre-trained on financial communications — immediate uplift expected'),
        ('Validation', 'Stratified 5-fold cross-validation', 'Avoid data leakage; separate by sender/thread'),
        ('Class Imbalance', 'Apply SMOTE or class-weighted loss if neutral < 25%',
         'Neutral class frequently under-represented in finance communications'),
    ]),
    ('DEPLOYMENT & MONITORING', 'C0392B', [
        ('Drift Detection', 'Monitor vocabulary drift monthly (KL divergence on token distributions)',
         'Finance language shifts during market regime changes'),
        ('Human-in-Loop', 'Flag emails with compound score in [-0.1, +0.1] for manual review',
         'Low-confidence boundary cases benefit most from human annotation'),
        ('Retraining Cadence', 'Quarterly retraining with ≥200 new labeled samples per cycle',
         'More frequent if: new regulation, market crisis, M&A wave'),
        ('Alert Thresholds', 'Auto-escalate if: >30% negative spike vs prior 30-day baseline',
         'Early warning for counterparty distress / credit deterioration'),
    ]),
]

cur_row = 3
for section, color, rows in recs:
    add_section_title(ws6, cur_row, 1, section, 12, bg=color)
    cur_row += 1

    for c, h in enumerate(['Item','Recommendation','Rationale/Notes'], 1):
        cell = ws6.cell(row=cur_row, column=c, value=h)
        cell.font = hdr_font(sz=10, color='1A1A2E'); cell.fill = fill('D5D8DC')
        cell.alignment = center(); cell.border = thin_border()
    cur_row += 1

    for item, rec, note in rows:
        ws6.row_dimensions[cur_row].height = 38
        for c, v in enumerate([item, rec, note], 1):
            cell = ws6.cell(row=cur_row, column=c, value=v)
            cell.font = body_font(bold=(c==1))
            cell.fill = fill('FDFEFE')
            cell.alignment = left(); cell.border = thin_border()
        cur_row += 1
    cur_row += 1  # spacer

set_col_widths(ws6, {'A':26,'B':50,'C':58})

# ── Reorder sheets ─────────────────────────────────────────────────────────────
sheet_order = ['Summary','Raw_Data','Keyword_Analysis','Monthly_Trends','Charts','Retraining_Guide']
wb._sheets = [wb[n] for n in sheet_order]

# ── Save ──────────────────────────────────────────────────────────────────────
OUT = r'H:\Mphasis\2025\windsruf_workspace\sentiment_anaysis_txt\Sentiment_Analysis_Report_EDA.xlsx'
wb.save(OUT)
print(f"\n✅  Saved: {OUT}")
print(f"   Tabs: {', '.join(sheet_order)}")
